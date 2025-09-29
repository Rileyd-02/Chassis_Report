import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill
import plotly.express as px

st.set_page_config(page_title="Subchassis Mapper", layout="wide")

st.title("📊 Subchassis Mapper Tool")
st.markdown("""
Upload your **Planning file** and **Subchassis reference report**.  
You will be guided step by step to select the correct columns for mapping.
""")

# --- Step 1: Upload Planning File ---
uploaded_planning = st.file_uploader("Upload Planning Excel File", type=["xlsx"])
planning_df = None
style_col_plan = None

if uploaded_planning:
    planning_excel = pd.ExcelFile(uploaded_planning, engine="openpyxl")
    sheet_names = planning_excel.sheet_names
    selected_sheet = st.selectbox("Select Sheet from Planning File", sheet_names)
    planning_df = planning_excel.parse(selected_sheet)
    st.success(f"✅ Loaded planning sheet: {selected_sheet}")

    # Let user pick style column from planning file (columns containing "style")
    style_candidates = [c for c in planning_df.columns if "style" in c.lower()]
    style_col_plan = st.selectbox(
        "Select Style Column in Planning File",
        style_candidates if style_candidates else planning_df.columns
    )

# --- Step 2: Upload Subchassis Reference File ---
uploaded_sub = st.file_uploader("Upload Subchassis Reference File", type=["xlsx"])
sub_df = None
style_col_sub = None
customer_col = None
dept_col = None
season_col = None

if uploaded_sub:
    sub_excel = pd.ExcelFile(uploaded_sub, engine="openpyxl")
    sheet_names_sub = sub_excel.sheet_names
    selected_sheet_sub = st.selectbox("Select Sheet from Subchassis File", sheet_names_sub)
    sub_df = sub_excel.parse(selected_sheet_sub)
    st.success(f"✅ Loaded subchassis sheet: {selected_sheet_sub}")

    # Pick style, customer, department, season columns
    style_candidates_sub = [c for c in sub_df.columns if "style" in c.lower()]
    style_col_sub = st.selectbox(
        "Select Style Column in Subchassis File",
        style_candidates_sub if style_candidates_sub else sub_df.columns
    )
    customer_col = st.selectbox("Select Customer Column", sub_df.columns)
    dept_col = st.selectbox("Select Department Column", sub_df.columns)
    season_col = st.selectbox("Select Season Column (Optional)", ["<None>"] + list(sub_df.columns))

# --- Step 3: Process Mapping ---
if planning_df is not None and sub_df is not None and st.button("Map Subchassis"):
    try:
        # Clean style columns
        planning_df[style_col_plan] = planning_df[style_col_plan].astype(str).str.strip()
        sub_df[style_col_sub] = sub_df[style_col_sub].astype(str).str.strip()

        # Always include style + latestsubchassis + customer + department (+ season if selected)
        merge_cols = [style_col_sub, "LatestSubChassis", customer_col, dept_col]
        if season_col and season_col != "<None>":
            merge_cols.append(season_col)

        # Merge
        merged_df = pd.merge(
            planning_df,
            sub_df[merge_cols],
            left_on=style_col_plan,
            right_on=style_col_sub,
            how="left"
        )

        # Drop duplicate key column if different
        if style_col_sub != style_col_plan:
            merged_df.drop(columns=[style_col_sub], inplace=True)

        # --- Filtering ---
        cust_options = sub_df[customer_col].dropna().unique().tolist()
        dept_options = sub_df[dept_col].dropna().unique().tolist()
        season_options = sub_df[season_col].dropna().unique().tolist() if season_col != "<None>" else []

        selected_customer = st.multiselect("Filter by Customer", cust_options)
        selected_department = st.multiselect("Filter by Department", dept_options)
        selected_season = st.multiselect("Filter by Season", season_options) if season_options else []

        filtered_df = merged_df.copy()
        if selected_customer:
            filtered_df = filtered_df[filtered_df[customer_col].isin(selected_customer)]
        if selected_department:
            filtered_df = filtered_df[filtered_df[dept_col].isin(selected_department)]
        if selected_season and season_col != "<None>":
            filtered_df = filtered_df[filtered_df[season_col].isin(selected_season)]

        # --- Summary Box ---
        total_rows = len(filtered_df)
        unique_styles = filtered_df[style_col_plan].nunique()
        missing_subchassis = filtered_df["LatestSubChassis"].isna().sum()

        st.markdown("### 📊 Filtered Summary")
        st.info(f"""
        - **Total Rows:** {total_rows}  
        - **Unique Styles:** {unique_styles}  
        - **Missing Subchassis:** {missing_subchassis}  
        """)

        # --- Chart ---
        if total_rows > 0:
            st.markdown("### 📈 Distribution Chart")
            chart_dim = st.radio("Select field to visualize", [customer_col, dept_col] + ([season_col] if season_col != "<None>" else []))
            fig = px.bar(
                filtered_df.groupby(chart_dim).size().reset_index(name="Count"),
                x=chart_dim,
                y="Count",
                title=f"Distribution by {chart_dim}",
                text="Count"
            )
            fig.update_traces(textposition="outside")
            st.plotly_chart(fig, use_container_width=True)

        # --- Preview ---
        st.markdown("### 🔍 Preview of Filtered Data")
        st.dataframe(filtered_df.head(20))

        # Save with highlights for missing LatestSubChassis
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            filtered_df.to_excel(writer, index=False, sheet_name="Mapped Data")
            ws = writer.sheets["Mapped Data"]

            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            latest_col_idx = filtered_df.columns.get_loc("LatestSubChassis") + 1
            for row_idx, value in enumerate(filtered_df["LatestSubChassis"], start=2):
                if pd.isna(value):
                    ws.cell(row=row_idx, column=latest_col_idx).fill = red_fill

        st.download_button(
            label="Download Filtered Excel File",
            data=output.getvalue(),
            file_name="Filtered_Mapped_Planning_Sheet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ An error occurred: {e}")
