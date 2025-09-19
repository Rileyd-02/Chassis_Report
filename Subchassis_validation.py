import streamlit as st
import pandas as pd
import difflib
from io import BytesIO
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Subchassis Mapper", layout="wide")

st.title("Subchassis Mapper Tool")
st.markdown("""
This tool lets you upload a dynamic planning Excel file and a standard Subchassis reference table.
It will map the 'LatestSubChassis' from the reference table to your planning sheet based on 'Style' and optionally 'Customer Department'.

**Instructions:**
1. Upload your dynamic planning Excel file and the Subchassis reference table.
2. The tool will try to identify 'Style' and 'Customer Department' columns using fuzzy matching.
3. You can confirm or adjust the detected columns.
4. The tool will perform a join and append the 'LatestSubChassis' column to your planning sheet.
5. Download the compiled Excel file with visual highlights for missing mappings.
""")

uploaded_dynamic = st.file_uploader("Upload Dynamic Planning Excel File", type=["xlsx"])
uploaded_reference = st.file_uploader("Upload Subchassis Reference Excel File", type=["xlsx"])

def fuzzy_match_column(possible_names, columns):
    matches = {}
    for name in possible_names:
        match = difflib.get_close_matches(name, columns, n=1, cutoff=0.6)
        if match:
            matches[name] = match[0]
    return list(set(matches.values()))

if uploaded_dynamic and uploaded_reference:
    try:
        # Load dynamic file
        dynamic_excel = pd.ExcelFile(uploaded_dynamic, engine='openpyxl')
        sheet_names = dynamic_excel.sheet_names
        selected_sheet = sheet_names[0] if len(sheet_names) == 1 else st.selectbox("Select Sheet from Dynamic File", sheet_names)
        dynamic_df = dynamic_excel.parse(selected_sheet)
        st.success(f"Loaded dynamic sheet: {selected_sheet}")

        # Load reference file
        reference_df = pd.read_excel(uploaded_reference, sheet_name=0, engine='openpyxl')
        st.success(f"Reference file loaded successfully")

        # Fuzzy match for Style and Customer Department
        dynamic_columns = dynamic_df.columns.astype(str).tolist()
        style_keywords = ['Style', 'Style #', 'Style No', 'Style number']
        customer_keywords = ['Customer Department', 'Department', 'Buying Office', 'Customer']

        matched_style = fuzzy_match_column(style_keywords, dynamic_columns)
        matched_customer = fuzzy_match_column(customer_keywords, dynamic_columns)

        style_col = st.selectbox("Select Style Column", matched_style if matched_style else dynamic_columns)
        customer_col = st.selectbox("Select Customer Department Column (Optional)", matched_customer + ['None'] if matched_customer else dynamic_columns + ['None'])

        if st.button("Map Subchassis"):
            st.info("Mapping subchassis, please wait...")

            # Clean string columns
            dynamic_df[style_col] = dynamic_df[style_col].astype(str).str.strip()
            reference_df['Style'] = reference_df['Style'].astype(str).str.strip()

            # Prepare join key
            dynamic_df['join_key'] = dynamic_df[style_col]
            reference_df['join_key'] = reference_df['Style']
            if customer_col != 'None':
                dynamic_df[customer_col] = dynamic_df[customer_col].astype(str).str.strip()
                reference_df['Department'] = reference_df['Department'].astype(str).str.strip()
                dynamic_df['join_key'] += "_" + dynamic_df[customer_col]
                reference_df['join_key'] += "_" + reference_df['Department']

            # Merge data
            merged_df = pd.merge(dynamic_df, reference_df[['join_key', 'LatestSubChassis']], on='join_key', how='left')
            merged_df.drop(columns=['join_key'], inplace=True)

            # Reorder columns to place LatestSubChassis at the end
            cols = [col for col in merged_df.columns if col != 'LatestSubChassis'] + ['LatestSubChassis']
            merged_df = merged_df[cols]

            # Save to Excel with highlights
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, sheet_name='Mapped Data')
                ws = writer.sheets['Mapped Data']

                # Apply red fill for missing LatestSubChassis
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                latest_col_idx = merged_df.columns.get_loc('LatestSubChassis') + 1  # Excel is 1-indexed
                for row_idx, value in enumerate(merged_df['LatestSubChassis'], start=2):
                    if pd.isna(value):
                        ws.cell(row=row_idx, column=latest_col_idx).fill = red_fill

            st.success("Mapping complete!")
            st.download_button(
                label="Download Mapped Excel File",
                data=output.getvalue(),
                file_name="Mapped_Planning_Sheet.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"An error occurred: {e}")
